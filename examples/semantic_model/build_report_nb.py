"""Multi-source test: semantic model + PDF + CSV + a custom skill -> formatted xlsx.

The point is not that an Excel file appears. It is that four sources each carry
facts the others cannot supply, so every wrong cell is attributable to a source
the agent failed to read:

  semantic model  the actual KPI values (trailing 30 days)
  targets.csv     the target for each KPI, and the owner
  memo PDF        three risk items, and the escalation threshold
  report_context  which measure answers which KPI, the 30-day window, the
                  direction of "good" per KPI, and the formatting conventions

Miss the CSV and Target/Variance are empty. Miss the PDF and Escalate and the
Notes sheet are wrong. Miss the skill and the agent picks `sls_amt_x` or
`Day Yield Pct` (both wrong, both plausible) and reports all-time figures.

Grading is post-hoc against ground truth computed here with sempy and pandas,
plus openpyxl formatting checks. `output_validator` only enforces that a
readable workbook was actually written - the mechanical floor, not the grade.
"""
import json
import pathlib

import nbformat as nbf

HERE = pathlib.Path(__file__).parent
SKILL = (HERE / "skills" / "report_context.md").read_text(encoding="utf-8")

# ---- change these three for your own workspace -------------------------
WORKSPACE_ID = "002964bb-c154-4279-a405-cac05ecb54a6"
LAKEHOUSE_ID = "d2317250-2392-49cb-8db3-e4327daa10bb"
LAKEHOUSE_NAME = "evalresults"
# The semantic model name is set in the notebook's parameter cell (MODEL_NAME)
# so it can be overridden per run with `fab job run -P MODEL_NAME:string=...`.
# ------------------------------------------------------------------------

BASE = "/lakehouse/default/Files/multisource"
OUT_XLSX = f"{BASE}/reports/ops_review.xlsx"

MD = """# Multi-source report generation

Semantic model + PDF + CSV + a custom context skill, into one formatted
workbook written to the lakehouse.

| source | supplies | how a miss shows up |
| --- | --- | --- |
| semantic model | actual KPI values, trailing 30 days | Actual column wrong |
| `targets.csv` | target and owner per KPI | Target/Variance/Owner empty |
| `ops_memo.pdf` | three risks, escalation threshold | Escalate wrong, Notes sheet wrong |
| `report_context` skill | which measure, which window, direction of good, formatting | wrong measure, all-time figures, unformatted |

Set `ARM` to `full` (all four) or `noskill` (drop the custom skill) to see what
the skill is carrying.
"""

CELL_PARAM = '''ARM = "full"             # "full" | "noskill"
MODEL_NAME = "Manufacturing Ops AI Ready"
MAX_TURNS = 30
TIMEOUT_S = 1200
print("arm:", ARM)
'''

CELL_INSTALL = (
    "%pip install -q reportlab pypdf openpyxl "
    "git+https://github.com/pawarbi/fabric-rlm-core.git@feat/semantic-model-input\n"
)

# --- fixtures -------------------------------------------------------------
TARGETS_ROWS = [
    ("Total Sales", 900000000.0, "currency", "R. Alvarez"),
    ("Production Yield", 0.98, "rate", "M. Chen"),
    ("Downtime", 0.03, "rate", "M. Chen"),
    ("Scrap Rate", 0.02, "rate", "K. Osei"),
]

RISKS = [
    ("R-101", "Rheinland furnace line 3 is scheduled for relining in August; "
              "expect sustained yield pressure through the outage."),
    ("R-102", "Grade-A billet remains single-sourced. No alternate supplier has "
              "completed qualification."),
    ("R-103", "Riverside night shift staffing is running 12 percent below plan."),
]

ESCALATION = (
    "Escalation rule: a KPI is escalated to the plant director when its variance "
    "is unfavourable by more than 0.20 percentage points for rate KPIs, or by "
    "more than 5 percent of target for currency KPIs. Favourable variances are "
    "never escalated."
)

CELL_FIXTURES = f'''import os, csv
BASE = "{BASE}"
os.makedirs(f"{{BASE}}/reports", exist_ok=True)

TARGETS = {TARGETS_ROWS!r}
CSV_PATH = f"{{BASE}}/targets.csv"
with open(CSV_PATH, "w", newline="", encoding="utf-8") as fh:
    w = csv.writer(fh)
    w.writerow(["kpi", "target", "kind", "owner"])
    for row in TARGETS:
        w.writerow(row)
print("wrote", CSV_PATH)

RISKS = {RISKS!r}
ESCALATION = {ESCALATION!r}

from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer

PDF_PATH = f"{{BASE}}/ops_memo.pdf"
styles = getSampleStyleSheet()
flow = [Paragraph("Q2 Operations Review - Preliminary Notes", styles["Title"]),
        Spacer(1, 10),
        Paragraph("Circulated to plant leadership ahead of the quarterly review. "
                  "Figures are produced separately from the operational model; "
                  "this memo carries the narrative items only.", styles["BodyText"]),
        Spacer(1, 10),
        Paragraph("Open risks", styles["Heading2"])]
for rid, text in RISKS:
    flow.append(Paragraph(f"<b>{{rid}}</b> - {{text}}", styles["BodyText"]))
    flow.append(Spacer(1, 6))
flow += [Spacer(1, 10), Paragraph("Escalation", styles["Heading2"]),
         Paragraph(ESCALATION, styles["BodyText"])]
SimpleDocTemplate(PDF_PATH, pagesize=LETTER).build(flow)
print("wrote", PDF_PATH)

from pypdf import PdfReader
_txt = "\\n".join(p.extract_text() or "" for p in PdfReader(PDF_PATH).pages)
assert "R-103" in _txt and "0.20 percentage points" in _txt, "PDF text is not extractable"
print("pdf text extractable:", len(_txt), "chars")
'''

CELL_SKILL = (
    'import os\n'
    f'SKILL_DIR = "{BASE}/skills"\n'
    'os.makedirs(SKILL_DIR, exist_ok=True)\n'
    'SKILL_TEXT = r"""' + SKILL + '"""\n'
    'with open(f"{SKILL_DIR}/report_context.md", "w", encoding="utf-8") as fh:\n'
    '    fh.write(SKILL_TEXT)\n'
    'print("report_context:", len(SKILL_TEXT), "chars")\n'
)

# --- ground truth ---------------------------------------------------------
CELL_TRUTH = f'''import sempy.fabric as fabric
import pandas as pd

DATASET = MODEL_NAME

def dax(q):
    return fabric.evaluate_dax(DATASET, q)

# Trailing 30 days ending at the model's last date - the window the skill
# defines, and the one every "Actual" must use.
_w = dax("EVALUATE ROW(\\"last\\", MAX('Date'[Date]))")
LAST = pd.to_datetime(_w.iloc[0, 0])
FROM = LAST - pd.Timedelta(days=29)
print("window:", FROM.date(), "->", LAST.date())

_win = "'Date'[Date] >= DATE({{}},{{}},{{}}) && 'Date'[Date] <= DATE({{}},{{}},{{}})".format(
    FROM.year, FROM.month, FROM.day, LAST.year, LAST.month, LAST.day)

_q = f"""
EVALUATE
CALCULATETABLE(
    ROW(
        "Total Sales", [Total Sales],
        "Production Yield", [Production Yield %],
        "Downtime", [Downtime %],
        "Scrap Rate", DIVIDE(SUM(ProductionLog[Scrap]), SUM(ProductionLog[Qty]))
    ),
    {{_win}}
)
"""
_a = dax(_q)
ACTUAL = {{str(c).split("[")[-1].strip("]"): float(v)
          for c, v in zip(_a.columns, _a.iloc[0])}}
print("actuals:", {{k: round(v, 6) for k, v in ACTUAL.items()}})
assert set(ACTUAL) == {{"Total Sales", "Production Yield", "Downtime", "Scrap Rate"}}, \\
    f"unexpected DAX column names: {{list(_a.columns)}}"

# Plant column is discovered rather than assumed - column naming varies.
_cols = fabric.list_columns(DATASET)
_cand = [(r["Table Name"], r["Column Name"]) for _, r in _cols.iterrows()
         if "plant" in str(r["Column Name"]).lower()]
print("plant column candidates:", _cand[:5])
assert _cand, ("no column with 'plant' in its name; ground truth cannot be "
               f"computed. Columns: {{list(_cols['Column Name'])[:40]}}")
PLANT_TBL, PLANT_COL = _cand[0]
_p = dax(f"""
EVALUATE
CALCULATETABLE(
    SUMMARIZECOLUMNS(
        '{{PLANT_TBL}}'[{{PLANT_COL}}],
        "Production Yield", [Production Yield %],
        "Downtime", [Downtime %],
        "Scrap Rate", DIVIDE(SUM(ProductionLog[Scrap]), SUM(ProductionLog[Qty]))
    ),
    {{_win}}
)
""")
_p.columns = [str(c).split("[")[-1].strip("]") for c in _p.columns]
BY_PLANT = {{str(r.iloc[0]): {{k: float(r[k]) for k in
            ("Production Yield", "Downtime", "Scrap Rate")}}
            for _, r in _p.iterrows()}}
print("by plant:", {{p: {{k: round(v, 5) for k, v in d.items()}}
                   for p, d in BY_PLANT.items()}})

TARGET = {{k: t for k, t, _kind, _own in TARGETS}}
KIND = {{k: kind for k, _t, kind, _own in TARGETS}}
OWNER = {{k: o for k, _t, _kind, o in TARGETS}}
HIGHER_IS_BETTER = {{"Total Sales": True, "Production Yield": True,
                    "Downtime": False, "Scrap Rate": False}}

GT = {{}}
for k, act in ACTUAL.items():
    tgt = TARGET[k]
    var = act - tgt
    favourable = (var >= 0) if HIGHER_IS_BETTER[k] else (var <= 0)
    if favourable:
        esc = False
    elif KIND[k] == "rate":
        esc = abs(var) > 0.0020          # 0.20 percentage points
    else:
        esc = abs(var) > 0.05 * tgt      # 5 percent of target
    GT[k] = dict(actual=act, target=tgt, variance=var,
                 favourable=favourable, escalate=esc, owner=OWNER[k])

for k, v in GT.items():
    print(f"  {{k:18s}} act={{v['actual']:>16,.4f}} tgt={{v['target']:>14,.4f}} "
          f"var={{v['variance']:>+12,.4f}} {{'OK ' if v['favourable'] else 'BELOW'}} "
          f"escalate={{v['escalate']}}")
'''

# --- the task -------------------------------------------------------------
CELL_TASK = f'''TASK = """Produce an Excel operations review and save it to the lakehouse.

Data sources, all four of which you need:

1. The semantic model bound as `model` in your namespace - the actual
   KPI values.
2. targets.csv at {{csv}} - the target and owner for each KPI.
3. ops_memo.pdf at {{pdf}} - narrative risk items and the escalation rule.
   Extract its text with pypdf, which is installed.
4. Your loaded skills - house conventions for which measure answers which KPI,
   the reporting window, and the workbook formatting rules.

Write the workbook to exactly this path:

    {OUT_XLSX}

It must contain three sheets, named exactly:

**Summary** - one row per KPI, in this order: Total Sales, Production Yield,
Downtime, Scrap Rate. Columns, in this order:

    KPI | Actual | Target | Variance | Status | Escalate | Owner

Status is "On Track" when the variance is favourable for that KPI and "Below
Target" when it is not. Remember that for Downtime and Scrap Rate a lower number
is better. Escalate is "Yes" or "No", decided by the rule in the memo. Owner
comes from the CSV.

**By Plant** - one row per plant, columns: Plant | Production Yield | Downtime |
Scrap Rate.

**Notes** - the open risks from the memo, columns: ID | Risk. One row per risk,
using the memo's own identifiers.

Also state the reporting window somewhere in the Summary sheet so a reader knows
what period the figures cover.

Apply the workbook formatting conventions from your skills.

When the file is written, SUBMIT with `path` set to the path you wrote and
`summary` describing what you produced."""


def build_task():
    return TASK.format(model=MODEL_NAME,
                       csv=f"{BASE}/targets.csv",
                       pdf=f"{BASE}/ops_memo.pdf")


print(build_task()[:400], "...")
'''

CELL_RUN = f'''import time, os, json
from fabric_rlm import RLM, SemanticModel, SkillLoader

with open("/lakehouse/default/Files/orkey.txt", encoding="utf-8") as fh:
    _key = fh.read().strip()
LM = {{"model": "openrouter/minimax/minimax-m3", "api_key": _key,
      "api_base": "https://openrouter.ai/api/v1", "timeout": 900}}

OUT = "{OUT_XLSX}"
if os.path.exists(OUT):
    os.remove(OUT)          # never grade a stale file from a previous run


def validate(payload):
    """Mechanical floor: a readable workbook exists where it was asked for.

    Deliberately not a grade - it checks nothing about values or formatting,
    so it cannot manufacture the result the experiment is measuring."""
    import openpyxl
    if not os.path.exists(OUT):
        raise ValueError(f"No workbook at {{OUT}}. Write the file before SUBMIT.")
    try:
        wb = openpyxl.load_workbook(OUT)
    except Exception as e:
        raise ValueError(f"Workbook at {{OUT}} does not open: {{type(e).__name__}}: {{e}}")
    missing = [s for s in ("Summary", "By Plant", "Notes") if s not in wb.sheetnames]
    if missing:
        raise ValueError(f"Workbook is missing sheet(s): {{missing}}. Found: {{wb.sheetnames}}")


# The semantic_model skill is not loaded here: the bound handle supplies the
# entry point it used to describe, and measured the same with or without it.
# report_context stays, because house conventions cannot come from an API.
skills = ["report_context"] if ARM == "full" else []
loader = SkillLoader("{BASE}/skills")

t0 = time.time()
r = RLM.task(
    task=build_task(),
    inputs={{"targets_csv": f"{BASE}/targets.csv",
            "ops_memo_pdf": f"{BASE}/ops_memo.pdf",
            "model": SemanticModel(MODEL_NAME),
            "output_path": OUT}},
    outputs=["path", "summary"],
    lm=LM,
    skills=skills,
    skill_loader=loader,
    output_validator=validate,
    max_turns=MAX_TURNS,
    timeout=TIMEOUT_S,
).run()
ELAPSED = round(time.time() - t0, 1)
REPORT = r.report(as_dict=True)
print("submitted:", r.submitted, "| turns:", REPORT["turns"],
      "| repairs:", REPORT.get("repair_turns"), "| secs:", ELAPSED)
print("answer:", str((r.payload or {{}}).get("summary", ""))[:400])
'''

CELL_GRADE = f'''import openpyxl, re, json

OUT = "{OUT_XLSX}"
checks = []            # (source, label, ok, detail)


def chk(source, label, ok, detail=""):
    checks.append(dict(source=source, label=label, ok=bool(ok), detail=str(detail)[:160]))


def near(a, b, tol=0.002):
    """Tight by default, deliberately.

    Rates here sit near 0.98, and the wrong reporting window moves them by
    ~0.1% relative. A 1% tolerance passed a Production Yield of 0.976613 when
    the right answer was 0.977740 - a false pass on the exact error the test
    exists to catch. Anything looser than ~0.2% cannot see a window mistake."""
    try:
        a, b = float(a), float(b)
    except (TypeError, ValueError):
        return False
    return abs(a - b) <= max(abs(b) * tol, 1e-9)


try:
    wb = openpyxl.load_workbook(OUT)
except Exception as e:
    wb = None
    chk("structure", "workbook opens", False, e)

if wb is not None:
    chk("structure", "workbook opens", True)
    for s in ("Summary", "By Plant", "Notes"):
        chk("structure", f"sheet {{s!r}}", s in wb.sheetnames, wb.sheetnames)

if wb is not None and "Summary" in wb.sheetnames:
    ws = wb["Summary"]
    grid = [[c.value for c in row] for row in ws.iter_rows()]

    def find_row(needle):
        n = re.sub(r"[^a-z]", "", str(needle).lower())
        for i, row in enumerate(grid):
            for v in row[:2]:
                if v is not None and re.sub(r"[^a-z]", "", str(v).lower()) == n:
                    return i
        return None

    def cells(i):
        return [c for c in grid[i]]

    for kpi, gt in GT.items():
        ri = find_row(kpi)
        if ri is None:
            chk("semantic model", f"{{kpi}}: row present", False, "not found")
            continue
        row = cells(ri)
        vals = [v for v in row if isinstance(v, (int, float))]
        strs = " ".join(str(v) for v in row if isinstance(v, str))

        chk("semantic model", f"{{kpi}}: actual",
            any(near(v, gt["actual"]) for v in vals), f"row={{row}}")
        chk("targets.csv", f"{{kpi}}: target",
            any(near(v, gt["target"]) for v in vals), f"row={{row}}")
        chk("report_context", f"{{kpi}}: variance",
            any(near(v, gt["variance"], 0.02) for v in vals), f"row={{row}}")
        want_status = "on track" if gt["favourable"] else "below"
        chk("report_context", f"{{kpi}}: status direction",
            want_status in strs.lower(), strs[:80])
        want_esc = "yes" if gt["escalate"] else "no"
        esc_cell = next((str(v).strip().lower() for v in row
                         if str(v).strip().lower() in ("yes", "no")), None)
        chk("ops_memo.pdf", f"{{kpi}}: escalate={{want_esc}}",
            esc_cell == want_esc, f"got {{esc_cell!r}}")
        chk("targets.csv", f"{{kpi}}: owner",
            gt["owner"].split()[-1].lower() in strs.lower(), strs[:80])

    # reporting window stated somewhere on the sheet
    flat = " ".join(str(v) for row in grid for v in row if v is not None)
    chk("report_context", "reporting window stated",
        str(LAST.year) in flat and (str(FROM.day) in flat or FROM.strftime("%b") in flat
                                    or str(FROM.date()) in flat), flat[:120])

    # formatting
    hdr = next((i for i, row in enumerate(grid)
                if row and any(isinstance(v, str) and v.strip().lower() == "kpi"
                               for v in row)), 0)
    hcells = [c for c in ws[hdr + 1] if c.value is not None]
    chk("report_context", "header bold", bool(hcells) and all(c.font.bold for c in hcells))
    chk("report_context", "header filled",
        bool(hcells) and any(c.fill is not None and c.fill.fgColor is not None
                             and str(c.fill.fgColor.rgb) not in ("00000000", "None")
                             for c in hcells))
    chk("report_context", "panes frozen", ws.freeze_panes is not None, ws.freeze_panes)
    chk("report_context", "column widths set",
        any(d.width for d in ws.column_dimensions.values()))

    fmts = [c.number_format for row in ws.iter_rows() for c in row
            if isinstance(c.value, (int, float))]
    chk("report_context", "currency format used",
        any("$" in f or "#,##0" in f for f in fmts), fmts[:6])
    chk("report_context", "percent format used",
        any("%" in f for f in fmts), fmts[:6])
    chk("report_context", "rates stored as fractions, not strings",
        not re.search(r"\\d+\\.\\d+%", " ".join(str(v) for row in grid for v in row
                                              if isinstance(v, str))))

if wb is not None and "By Plant" in wb.sheetnames:
    ws = wb["By Plant"]
    grid = [[c.value for c in row] for row in ws.iter_rows()]
    flat = " ".join(str(v) for row in grid for v in row if v is not None).lower()
    for plant, vals in BY_PLANT.items():
        chk("semantic model", f"plant {{plant}}: present", plant.lower() in flat)
        ri = next((i for i, row in enumerate(grid)
                   if any(str(v).strip().lower() == plant.lower() for v in row
                          if v is not None)), None)
        if ri is None:
            continue
        nums = [v for v in grid[ri] if isinstance(v, (int, float))]
        for metric, want in vals.items():
            chk("semantic model", f"plant {{plant}}: {{metric}}",
                any(near(v, want) for v in nums), f"nums={{[round(n,5) for n in nums]}}")

if wb is not None and "Notes" in wb.sheetnames:
    ws = wb["Notes"]
    flat = " ".join(str(c.value) for row in ws.iter_rows() for c in row
                    if c.value is not None)
    for rid, text in RISKS:
        chk("ops_memo.pdf", f"note {{rid}} id", rid in flat)
        key = text.split()[0].lower()
        chk("ops_memo.pdf", f"note {{rid}} text",
            any(w in flat.lower() for w in (key, text.split()[1].lower())))

passed = sum(1 for c in checks if c["ok"])
print()
for c in checks:
    if not c["ok"]:
        print(f"  MISS [{{c['source']:15s}}] {{c['label']}}  {{c['detail']}}")
print()
print(f"ARM={{ARM}}: {{passed}}/{{len(checks)}} checks passed")
by_src = {{}}
for c in checks:
    d = by_src.setdefault(c["source"], [0, 0])
    d[1] += 1
    d[0] += int(c["ok"])
for src, (p, n) in sorted(by_src.items()):
    print(f"   {{src:16s}} {{p:2d}}/{{n:2d}}")

payload = dict(arm=ARM, passed=passed, total=len(checks), by_source=by_src,
               turns=REPORT.get("turns"), repairs=REPORT.get("repair_turns"),
               submitted=bool(REPORT.get("submitted")), secs=ELAPSED,
               checks=checks)
with open(f"/lakehouse/default/Files/multisource_{{ARM}}.json", "w",
          encoding="utf-8") as fh:
    json.dump(payload, fh, indent=1, default=str)
print("wrote multisource_" + ARM + ".json")
'''


def param_cell(src):
    c = nbf.v4.new_code_cell(src)
    c.metadata["tags"] = ["parameters"]
    return c


nb = nbf.v4.new_notebook(cells=[
    nbf.v4.new_markdown_cell(MD),
    param_cell(CELL_PARAM),
    nbf.v4.new_code_cell(CELL_INSTALL),
    nbf.v4.new_code_cell(CELL_FIXTURES),
    nbf.v4.new_code_cell(CELL_SKILL),
    nbf.v4.new_code_cell(CELL_TRUTH),
    nbf.v4.new_code_cell(CELL_TASK),
    nbf.v4.new_code_cell(CELL_RUN),
    nbf.v4.new_code_cell(CELL_GRADE),
])
nb.metadata.update({
    "kernel_info": {"name": "jupyter", "jupyter_kernel_name": "python3.12"},
    "kernelspec": {"name": "jupyter", "display_name": "Jupyter"},
    "language_info": {"name": "python"},
    "microsoft": {"language": "python", "language_group": "jupyter_python"},
    "dependencies": {"lakehouse": {
        "default_lakehouse": LAKEHOUSE_ID,
        "default_lakehouse_name": LAKEHOUSE_NAME,
        "default_lakehouse_workspace_id": WORKSPACE_ID}},
})
nbf.validate(nb)
_, nb = nbf.validator.normalize(nb)
for i, c in enumerate(nb.cells):
    if c.cell_type == "code" and not c.source.strip().startswith("%"):
        compile(c.source, f"<cell{i}>", "exec")

IT = HERE / "multisource_report.Notebook"
IT.mkdir(exist_ok=True)
(IT / ".platform").write_text(json.dumps({
    "$schema": "https://developer.microsoft.com/json-schemas/fabric/gitIntegration/"
               "platformProperties/2.0.0/schema.json",
    "metadata": {"type": "Notebook", "displayName": "multisource_report"},
    "config": {"version": "2.0", "logicalId": "00000000-0000-0000-0000-000000000000"},
}, indent=4), encoding="utf-8", newline="\n")
with open(IT / "notebook-content.ipynb", "w", encoding="utf-8", newline="\n") as fh:
    nbf.write(nb, fh)
print(f"built {IT.name}: {len(nb.cells)} cells, all compile")
