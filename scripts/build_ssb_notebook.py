"""Build a SpreadsheetBench Fabric notebook (one strategy per notebook).

Strategies:
  A: gpt-5 single-shot via dspy.Predict — produces python code, we exec it in subprocess.
  F: gpt-4.1-mini + fabric_rlm.RLM(engine='adaptive', EffortBanditPolicy)
     with skills=['data_exploration'] so the model writes & runs its own openpyxl code.

Grading: cell-by-cell exact match within answer_position on answer_sheet vs golden xlsx.

Usage:
  python scripts/build_ssb_notebook.py --strategy A --model gpt-5    --effort medium --run-id ssb-A-...   [--smoke 1]
  python scripts/build_ssb_notebook.py --strategy F --model gpt-4.1-mini --effort medium --run-id ssb-F-...
"""
import argparse, json, pathlib

ap = argparse.ArgumentParser()
ap.add_argument('--strategy', choices=['A','F'], required=True)
ap.add_argument('--model', default='gpt-4.1-mini')
ap.add_argument('--effort', default='medium')
ap.add_argument('--run-id', required=True)
ap.add_argument('--smoke', type=int, default=0, help='if >0, only run first N questions')
ap.add_argument('--out', default=None)
args = ap.parse_args()

WHEEL_PATH = "/lakehouse/default/Files/fabric_rlm_longcot/wheels/fabric_rlm-0.2.1.dev2+excelskill-py3-none-any.whl"
DATASET_TAR = "/lakehouse/default/Files/fabric_rlm_longcot/datasets/ssb_subset_50.tar.gz"
RUN_ROOT = f"/lakehouse/default/Files/fabric_rlm_adaptive_validation/spreadsheetbench/{args.run_id}"

cells = []

def cell(src):
    cells.append({'cell_type': 'code', 'metadata': {}, 'source': src.splitlines(keepends=True),
                  'outputs': [], 'execution_count': None})

# ---- Cell 1: install ----
cell(f"""%pip install -q dspy>=3.0.4 openpyxl
%pip install -q "{WHEEL_PATH}"
""")

# ---- Cell 2: imports + dataset extract + grader ----
cell(f"""import os, sys, json, time, tarfile, shutil, subprocess, traceback, pathlib, re
import openpyxl
import dspy
import fabric_rlm
from fabric_rlm import RLM, FabricLM

STRATEGY = "{args.strategy}"
MODEL    = "{args.model}"
EFFORT   = "{args.effort}"
RUN_ID   = "{args.run_id}"
SMOKE_N  = {args.smoke}

DATASET_TAR = "{DATASET_TAR}"
RUN_ROOT    = pathlib.Path("{RUN_ROOT}")
RUN_ROOT.mkdir(parents=True, exist_ok=True)
RESULTS_PATH = RUN_ROOT / f"results_{{STRATEGY}}.jsonl"
SUMMARY_PATH = RUN_ROOT / f"summary_{{STRATEGY}}.json"
TRACES_DIR   = RUN_ROOT / f"traces_{{STRATEGY}}"
TRACES_DIR.mkdir(exist_ok=True)

def stage(event, **kw):
    print(json.dumps({{"t": round(time.time(),2), "event": event, **kw}}, default=str))

stage("imports_done", fabric_rlm=getattr(fabric_rlm, '__version__', '?'))

WORK = pathlib.Path("/tmp/ssb_work")
WORK.mkdir(parents=True, exist_ok=True)
DS_DIR = WORK / "ds"
if not DS_DIR.exists():
    DS_DIR.mkdir()
    with tarfile.open(DATASET_TAR) as tf:
        tf.extractall(DS_DIR)
n_xlsx = len(list(DS_DIR.rglob('*.xlsx')))
stage("dataset_extracted", path=str(DS_DIR), xlsx_count=n_xlsx)

records = [json.loads(l) for l in open(DS_DIR / "ssb_subset_50.jsonl", encoding='utf-8')]
if SMOKE_N > 0:
    records = records[:SMOKE_N]
stage("records_loaded", n=len(records))

def grade(out_xlsx, gold_xlsx, sheet, cell_range):
    try:
        wb_a = openpyxl.load_workbook(out_xlsx, data_only=True)
        wb_b = openpyxl.load_workbook(gold_xlsx, data_only=True)
    except Exception as e:
        return False, 0, 0, f"load_err: {{e}}"
    sname_a = sheet if (sheet and sheet in wb_a.sheetnames) else wb_a.sheetnames[0]
    sname_b = sheet if (sheet and sheet in wb_b.sheetnames) else wb_b.sheetnames[0]
    a = wb_a[sname_a]; b = wb_b[sname_b]
    try:
        rng_a = a[cell_range]; rng_b = b[cell_range]
    except Exception as e:
        return False, 0, 0, f"range_err: {{e}}"
    flat_a = [c.value for row in rng_a for c in row]
    flat_b = [c.value for row in rng_b for c in row]
    if len(flat_a) != len(flat_b):
        return False, 0, len(flat_b), f"len_mismatch {{len(flat_a)}} vs {{len(flat_b)}}"
    matches = 0
    for x, y in zip(flat_a, flat_b):
        if (x is None and y is None) or (str(x).strip() == str(y).strip()):
            matches += 1
    return matches == len(flat_a), matches, len(flat_a), None
""")

# ---- Cell 3: build LM ----
effort_kw = f', reasoning_effort="{args.effort}"' if args.model.startswith(('gpt-5','o3','o1')) else ''
cell(f"""os.environ["FABRIC_RLM_CAPTURE_TURNS"] = "1"
base_lm = FabricLM("{args.model}"{effort_kw}, max_tokens=16000)
stage("lm_built", model="{args.model}", effort="{args.effort}")
""")

# ---- Cell 4: per-strategy runner ----
if args.strategy == 'A':
    runner = '''dspy.configure(lm=base_lm)

class SsbWriteCode(dspy.Signature):
    """Given an Excel manipulation instruction and the path to a working .xlsx,
    write a complete Python program that uses openpyxl (and pandas/numpy if helpful)
    to perform the manipulation IN-PLACE and save the workbook back to the same path.

    CRITICAL: The grader reads cell values with openpyxl(data_only=True), which does
    NOT evaluate Excel formulas. You MUST write the COMPUTED NUMERIC/STRING VALUES
    into the target cells, not Excel formulas. Compute everything in Python and write
    the literal result.
    """
    instruction: str = dspy.InputField()
    xlsx_path: str   = dspy.InputField()
    sheet_hint: str  = dspy.InputField(desc="Target sheet name; may be empty")
    answer_range: str = dspy.InputField(desc="Cell range that the grader will inspect, e.g. A3:D32")
    code: str = dspy.OutputField(desc="ONLY a Python program inside a ```python``` fenced block. Must end by saving the workbook back to xlsx_path. Write computed values, NOT Excel formulas.")

predict = dspy.Predict(SsbWriteCode)

def extract_code(text):
    m = re.search(r"```(?:python)?\\s*\\n(.*?)```", text or "", re.DOTALL)
    return (m.group(1) if m else (text or "")).strip()

t_start = time.time()
n_pass = 0
with RESULTS_PATH.open("w", encoding="utf-8") as out_fh:
    for idx, rec in enumerate(records):
        qid = rec['question_id']; sid = str(rec['spreadsheet_id'])
        init_src = DS_DIR / 'spreadsheets' / sid / rec['init_file']
        gold_src = DS_DIR / 'spreadsheets' / sid / rec['golden_file']
        work_dir = WORK / qid; work_dir.mkdir(exist_ok=True)
        work_xlsx = work_dir / 'work.xlsx'
        result = {"strategy": STRATEGY, "question_id": qid, "spreadsheet_id": sid,
                  "instruction_type": rec['instruction_type'],
                  "answer_sheet": rec.get('answer_sheet'),
                  "answer_position": rec['answer_position']}
        try:
            shutil.copyfile(init_src, work_xlsx)
            t0 = time.perf_counter()
            pred = predict(instruction=rec['instruction'], xlsx_path=str(work_xlsx),
                           sheet_hint=rec.get('answer_sheet') or '',
                           answer_range=rec['answer_position'])
            elapsed_lm = time.perf_counter() - t0
            code = extract_code(pred.code)
            history = getattr(base_lm, "history", []) or []
            last = history[-1] if history else {}
            usage = (last.get("usage") or {}) if isinstance(last, dict) else {}
            prompt_tok = usage.get("prompt_tokens") or usage.get("input_tokens") or 0
            completion_tok = usage.get("completion_tokens") or usage.get("output_tokens") or 0
            # Run code in subprocess
            t1 = time.perf_counter()
            p = subprocess.run([sys.executable, "-c", code],
                               capture_output=True, text=True, timeout=180)
            elapsed_exec = time.perf_counter() - t1
            (TRACES_DIR / f"trace_{qid}.json").write_text(json.dumps({
                "qid": qid, "instruction": rec['instruction'],
                "code": code, "returncode": p.returncode,
                "stdout": p.stdout[-2000:], "stderr": p.stderr[-2000:],
            }, indent=2), encoding='utf-8')
            passed, m, n, gerr = grade(str(work_xlsx), str(gold_src),
                                       rec.get('answer_sheet') or '', rec['answer_position'])
            result.update({
                "passed": passed, "cells_matched": m, "cells_total": n,
                "grade_err": gerr,
                "lm_seconds": round(elapsed_lm,2), "exec_seconds": round(elapsed_exec,2),
                "subprocess_returncode": p.returncode,
                "stderr_tail": p.stderr[-300:],
                "prompt_tokens": prompt_tok, "completion_tokens": completion_tok,
            })
        except Exception as e:
            result.update({"passed": False, "error": repr(e),
                           "traceback": traceback.format_exc()[:500]})
        if result.get("passed"): n_pass += 1
        out_fh.write(json.dumps(result, default=str) + "\\n"); out_fh.flush()
        stage("q_done", idx=idx+1, qid=qid, passed=result.get("passed"),
              cells=f"{result.get('cells_matched',0)}/{result.get('cells_total',0)}",
              lm_s=result.get("lm_seconds"), exec_s=result.get("exec_seconds"))

summary = {"strategy": STRATEGY, "model": MODEL, "effort": EFFORT, "run_id": RUN_ID,
           "n": len(records), "n_passed": n_pass,
           "pass_rate": round(n_pass/max(1,len(records)), 4),
           "total_seconds": round(time.time()-t_start, 1)}
SUMMARY_PATH.write_text(json.dumps(summary, indent=2), encoding='utf-8')
print("\\n=== SUMMARY ===")
print(json.dumps(summary, indent=2))
'''
else:
    # Strategy F: small model (gpt-4.1-mini) with fabric_rlm skill (interpreter).
    # NOTE: bandit policy can't be used here — gpt-4.1-mini doesn't accept
    # reasoning_effort, so the EffortBandit's per-rung effort lever crashes.
    # Headline claim is "small model + subprocess access beats big model alone";
    # the interpreter is what matters, not the effort ladder.
    runner = '''os.environ["FABRIC_RLM_PVR_MODE"] = "full"
os.environ.pop("FABRIC_RLM_PVR", None)

t_start = time.time()
n_pass = 0
with RESULTS_PATH.open("w", encoding="utf-8") as out_fh:
    for idx, rec in enumerate(records):
        qid = rec['question_id']; sid = str(rec['spreadsheet_id'])
        init_src = DS_DIR / 'spreadsheets' / sid / rec['init_file']
        gold_src = DS_DIR / 'spreadsheets' / sid / rec['golden_file']
        work_dir = WORK / qid; work_dir.mkdir(exist_ok=True)
        work_xlsx = work_dir / 'work.xlsx'
        rec_out = {"strategy": STRATEGY, "question_id": qid, "spreadsheet_id": sid,
                   "instruction_type": rec['instruction_type'],
                   "answer_sheet": rec.get('answer_sheet'),
                   "answer_position": rec['answer_position']}
        try:
            shutil.copyfile(init_src, work_xlsx)
            sheet_for_grade = rec.get('answer_sheet') or ''
            answer_pos = rec['answer_position']
            sheet_str = rec.get('answer_sheet') or '(use the only sheet in the workbook)'
            prompt_text = (
                f"You must MODIFY an Excel (.xlsx) workbook in place using openpyxl. "
                f"This is NOT a log/JSON/CSV exploration task — it is an Excel manipulation task.\\n\\n"
                f"WORKBOOK PATH (real .xlsx — open, edit, save back to this same path):\\n  {str(work_xlsx)}\\n"
                f"TARGET SHEET: {sheet_str}\\n"
                f"TARGET CELL RANGE (the grader inspects ONLY this range): {answer_pos}\\n\\n"
                f"INSTRUCTION:\\n{rec['instruction']}\\n\\n"
                f"REQUIRED PROTOCOL:\\n"
                f"  1. `import openpyxl`; `wb = openpyxl.load_workbook(r'{str(work_xlsx)}')`; "
                f"`ws = wb[<target sheet>]` (or wb.active if sheet not named). Print headers + a few sample rows.\\n"
                f"  2. Compute every required answer in pure Python (use openpyxl to read source data, "
                f"plain Python/pandas to compute).\\n"
                f"  3. Write the COMPUTED LITERAL VALUES into each cell of TARGET CELL RANGE — "
                f"e.g. `ws['B3'] = 12345.67`. Do NOT write Excel formulas like `=SUM(...)`.\\n"
                f"  4. `wb.save(r'{str(work_xlsx)}')` — save back to the SAME path.\\n"
                f"  5. Verify: reload with `openpyxl.load_workbook(r'{str(work_xlsx)}', data_only=True)` "
                f"and print the values in TARGET CELL RANGE. Confirm none are None and none start with '='.\\n"
                f"  6. Reply with the single word `done` as your final answer.\\n\\n"
                f"HARD RULES:\\n"
                f"  - The grader uses `openpyxl(data_only=True)` which does NOT evaluate formulas. "
                f"If you write `=IF(...)` or any formula starting with `=`, the grader sees None and you fail.\\n"
                f"  - Your final `answer` field must be exactly `done`. Do not put a formula or explanation there.\\n"
                f"  - This is an Excel file. Do not try to read it as JSONL / CSV / log file."
            )
            rlm = RLM(signature="question -> answer", lm=base_lm,
                      engine="v6-custom",
                      max_turns=14, skills=["excel_modify"], timeout=300.0)
            t0 = time.perf_counter()
            rlm_result = rlm.run({"question": prompt_text})
            elapsed = time.perf_counter() - t0
            traj = rlm_result.trajectory
            turn_records = list(getattr(traj, "turns", []) or []) if traj is not None else []
            n_turns = len(turn_records)
            prompt_tok = sum((getattr(t, "prompt_tokens", None) or 0) for t in turn_records)
            completion_tok = sum((getattr(t, "completion_tokens", None) or 0) for t in turn_records)
            passed, m, n, gerr = grade(str(work_xlsx), str(gold_src), sheet_for_grade, answer_pos)
            (TRACES_DIR / f"trace_{qid}.json").write_text(json.dumps({
                "qid": qid, "prompt": prompt_text, "passed": passed,
                "submitted": rlm_result.submitted, "n_turns": n_turns,
                "answer": str((rlm_result.payload or {}).get("answer")) if rlm_result.payload else None,
                "turns": [t.to_dict() if hasattr(t, "to_dict") else t for t in turn_records],
            }, default=str, indent=2), encoding='utf-8')
            rec_out.update({
                "passed": passed, "cells_matched": m, "cells_total": n, "grade_err": gerr,
                "submitted": rlm_result.submitted,
                "elapsed_seconds": round(elapsed, 2),
                "n_turns": n_turns,
                "prompt_tokens": prompt_tok, "completion_tokens": completion_tok,
            })
        except Exception as e:
            rec_out.update({"passed": False, "error": repr(e),
                            "traceback": traceback.format_exc()[:500]})
        if rec_out.get("passed"): n_pass += 1
        out_fh.write(json.dumps(rec_out, default=str) + "\\n"); out_fh.flush()
        stage("q_done", idx=idx+1, qid=qid, passed=rec_out.get("passed"),
              cells=f"{rec_out.get('cells_matched',0)}/{rec_out.get('cells_total',0)}",
              elapsed=rec_out.get("elapsed_seconds"),
              n_turns=rec_out.get("n_turns"))

summary = {"strategy": STRATEGY, "model": MODEL, "effort": EFFORT, "run_id": RUN_ID,
           "n": len(records), "n_passed": n_pass,
           "pass_rate": round(n_pass/max(1,len(records)), 4),
           "total_seconds": round(time.time()-t_start, 1)}
SUMMARY_PATH.write_text(json.dumps(summary, indent=2), encoding='utf-8')
print("\\n=== SUMMARY ===")
print(json.dumps(summary, indent=2))
'''

cell(runner)

nb = {
    'nbformat': 4, 'nbformat_minor': 5,
    'metadata': {
        'kernelspec': {'display_name': 'Python 3.12', 'language': 'python', 'name': 'python3.12'},
        'language_info': {'name': 'python', 'version': '3.12'},
        'kernel_info': {'name': 'jupyter', 'jupyter_kernel_name': 'python3.12'},
        'microsoft': {'language': 'python', 'language_group': 'jupyter_python'},
        'dependencies': {'lakehouse': {
            'default_lakehouse': '9d10bce5-1edc-4875-83c4-ac0a98a02775',
            'default_lakehouse_name': 'diagnostic',
            'default_lakehouse_workspace_id': '82ad2591-974a-4ad4-ace6-e24879274a4b',
        }},
    },
    'cells': cells,
}

out = args.out or f"notebooks/ssb_{args.strategy}_{args.model.replace('-','_').replace('.','_')}{('_smoke'+str(args.smoke)) if args.smoke else ''}.ipynb"
out_path = pathlib.Path(__file__).parent.parent / out
out_path.parent.mkdir(parents=True, exist_ok=True)
out_path.write_text(json.dumps(nb, indent=1), encoding='utf-8')
print(f"wrote {out_path}  cells={len(cells)}  strategy={args.strategy} model={args.model} effort={args.effort} smoke={args.smoke}")
