"""Excel artifact helpers for RLM validators and workbook-editing skills.

These utilities are intentionally benchmark-agnostic: they parse Excel target
ranges, iterate saved workbook cells, and validate common artifact mistakes
without knowing any golden answers.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Any, Iterable


@dataclass(frozen=True)
class ExcelTargetRange:
    sheet_name: str | None
    cell_range: str


@dataclass(frozen=True)
class ExcelCellValue:
    sheet_name: str
    coordinate: str
    value: Any


ERROR_LITERALS: frozenset[str] = frozenset(
    {"#N/A", "#VALUE!", "#REF!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!"}
)
CODE_OR_PROSE_MARKERS: tuple[str, ...] = (
    "Sub ",
    "End Sub",
    "Power Query",
    "VBA",
    "Macro:",
    "let Source",
    "Application.",
    "ws.Range",
    "ws.Rows",
)
PLACEHOLDERS: frozenset[str] = frozenset({"-", "TBD", "N/A", "see notes"})


def parse_target_ranges(position: str) -> list[ExcelTargetRange]:
    """Parse a comma-separated Excel target-range expression.

    Supports bare ranges (``A1:B2``), sheet-qualified ranges
    (``Sheet2!A1:B2``), quoted sheet names with commas
    (``'Sheet, 3'!A1:B2``), and the common abbreviated single-column form
    (``H2:27`` -> ``H2:H27``).
    """

    parts = _split_range_list(position)
    if not parts:
        raise ValueError("target range is empty")
    return [_parse_one_range(part) for part in parts]


def iter_target_cells(
    workbook_path: str | Path,
    position: str,
    *,
    default_sheet: str | None = None,
    data_only: bool = True,
) -> Iterable[ExcelCellValue]:
    """Yield target cells from a saved workbook in row-major order."""

    openpyxl = _openpyxl()
    wb = openpyxl.load_workbook(workbook_path, data_only=data_only)
    for target in parse_target_ranges(position):
        sheet_name = _resolve_sheet(wb.sheetnames, target.sheet_name, default_sheet)
        ws = wb[sheet_name]
        for coord, value in _coord_values(ws[target.cell_range]):
            yield ExcelCellValue(sheet_name=sheet_name, coordinate=coord, value=value)


def validate_target_range_sanity(
    workbook_path: str | Path,
    position: str,
    *,
    default_sheet: str | None = None,
    error_literals: Iterable[str] = ERROR_LITERALS,
    code_or_prose_markers: Iterable[str] = CODE_OR_PROSE_MARKERS,
    placeholders: Iterable[str] = PLACEHOLDERS,
) -> None:
    """Reject common invalid Excel artifact states without golden answers.

    Intentional blanks are allowed. Formula strings, Excel error literals,
    placeholder text, and code/prose markers are rejected in the target cells.
    Raises ``AssertionError`` with a repair-friendly message on the first issue.
    """

    openpyxl = _openpyxl()
    wb_formula = openpyxl.load_workbook(workbook_path, data_only=False)
    wb_values = openpyxl.load_workbook(workbook_path, data_only=True)
    errors = set(error_literals)
    markers = tuple(code_or_prose_markers)
    placeholder_set = set(placeholders)

    for target in parse_target_ranges(position):
        sheet_name = _resolve_sheet(wb_formula.sheetnames, target.sheet_name, default_sheet)
        formula_cells = _coord_values(wb_formula[sheet_name][target.cell_range])
        value_cells = _coord_values(wb_values[sheet_name][target.cell_range])
        if len(formula_cells) != len(value_cells):
            raise AssertionError(
                f"{sheet_name}!{target.cell_range} has inconsistent cell counts "
                f"between formula/data_only views ({len(formula_cells)} vs {len(value_cells)})"
            )
        for (coord, formula_val), (_, data_val) in zip(formula_cells, value_cells):
            for val, mode in ((formula_val, "formula"), (data_val, "data_only")):
                if val in errors:
                    raise AssertionError(f"{sheet_name}!{coord} has Excel error {val!r} in {mode} view")
            if formula_val is None:
                continue
            if isinstance(formula_val, str):
                if formula_val.startswith("="):
                    raise AssertionError(f"{sheet_name}!{coord} still contains a formula string")
                if formula_val in placeholder_set:
                    raise AssertionError(f"{sheet_name}!{coord} contains placeholder {formula_val!r}")
                for marker in markers:
                    if marker in formula_val:
                        raise AssertionError(
                            f"{sheet_name}!{coord} contains code/prose marker {marker!r}"
                        )


def summarize_workbook_context(
    workbook_path: str | Path,
    *,
    target_position: str,
    default_sheet: str | None = None,
    max_sample_rows: int = 5,
    max_sample_cols: int = 12,
) -> str:
    """Return a compact, read-only workbook summary for RLM task prompts.

    The summary is intentionally small and evidence-oriented: sheet names,
    target ranges, dimensions, headers, formula/merged-cell counts, and a few
    data-only sample rows. It gives the model workbook context without dumping
    large sheets into the prompt.
    """

    openpyxl = _openpyxl()
    wb_formula = openpyxl.load_workbook(workbook_path, data_only=False)
    wb_values = openpyxl.load_workbook(workbook_path, data_only=True)
    targets = parse_target_ranges(target_position)
    resolved_targets, missing_sheets = _resolve_context_targets(
        wb_formula.sheetnames, targets, default_sheet
    )
    existing_target_sheets = [sheet for sheet, _, exists in resolved_targets if exists]
    primary_sheet = existing_target_sheets[0] if existing_target_sheets else wb_formula.sheetnames[0]
    ws_formula = wb_formula[primary_sheet]
    ws_values = wb_values[primary_sheet]
    shown_cols = min(max(ws_formula.max_column, _max_target_column(targets)), max_sample_cols)
    shown_rows = min(ws_formula.max_row, max(1, max_sample_rows + 1))

    lines = [
        "WORKBOOK_CONTEXT",
        f"sheets: {', '.join(wb_formula.sheetnames)}",
        "target_ranges: "
        + ", ".join(
            f"{sheet_name}!{target.cell_range}"
            for sheet_name, target, _ in resolved_targets
        ),
        *(
            [
                "target_sheet_status: "
                + "; ".join(f"{sheet} missing in current workbook" for sheet in missing_sheets)
            ]
            if missing_sheets
            else []
        ),
        f"active_sheet: {primary_sheet}",
        f"dimensions: {ws_formula.dimensions} rows={ws_formula.max_row} cols={ws_formula.max_column}",
        f"merged_ranges: {len(ws_formula.merged_cells.ranges)}",
        f"formulas: {_formula_count(ws_formula)}",
        "headers: " + _format_row(ws_values, 1, shown_cols, quote_strings=False),
    ]
    for row in range(2, shown_rows + 1):
        lines.append(f"row {row}: " + _format_row(ws_values, row, shown_cols, quote_strings=True))
    return "\n".join(lines)


def summarize_workbook_structure_context(
    workbook_path: str | Path,
    *,
    target_position: str,
    default_sheet: str | None = None,
    max_sample_cols: int = 12,
) -> str:
    """Return workbook structure only, omitting sample row values.

    This lower-risk prompt context is meant to orient the model to sheets,
    target ranges, dimensions, and headers while still forcing it to inspect
    task-specific row values in its own first code turn.
    """

    summary = summarize_workbook_context(
        workbook_path,
        target_position=target_position,
        default_sheet=default_sheet,
        max_sample_rows=0,
        max_sample_cols=max_sample_cols,
    )
    return summary.replace("WORKBOOK_CONTEXT", "WORKBOOK_STRUCTURE_CONTEXT", 1) + "\nsample_rows: omitted"


def add_excel_workbook_context(
    task: str,
    workbook_path: str | Path,
    *,
    target_position: str,
    default_sheet: str | None = None,
    mode: str = "structure",
    max_sample_rows: int = 5,
    max_sample_cols: int = 12,
) -> str:
    """Return ``task`` with an explicit, read-only workbook context block.

    This helper is opt-in and leaves ``RLM.from_task(...)`` defaults unchanged.
    Use ``mode="structure"`` for sheet/range/header context without row samples,
    or ``mode="full"`` when a small data-only sample is useful.
    """

    if mode == "structure":
        context = summarize_workbook_structure_context(
            workbook_path,
            target_position=target_position,
            default_sheet=default_sheet,
            max_sample_cols=max_sample_cols,
        )
    elif mode == "full":
        context = summarize_workbook_context(
            workbook_path,
            target_position=target_position,
            default_sheet=default_sheet,
            max_sample_rows=max_sample_rows,
            max_sample_cols=max_sample_cols,
        )
    else:
        raise ValueError("mode must be 'structure' or 'full'")
    return f"{context}\n\nTASK\n{task}"


def _split_range_list(position: str) -> list[str]:
    parts: list[str] = []
    current: list[str] = []
    in_quote = False
    text = position.strip()
    i = 0
    while i < len(text):
        ch = text[i]
        if ch == "'":
            if in_quote:
                if i + 1 < len(text) and text[i + 1] == "'":
                    current.extend(["'", "'"])
                    i += 2
                    continue
                in_quote = False
            elif not "".join(current).strip():
                in_quote = True
            current.append(ch)
            i += 1
            continue
        if ch == "," and not in_quote:
            part = "".join(current).strip()
            if part:
                parts.append(part)
            current = []
        else:
            current.append(ch)
        i += 1
    part = "".join(current).strip()
    if part:
        parts.append(part)
    return parts


def _parse_one_range(part: str) -> ExcelTargetRange:
    sheet_name: str | None = None
    cell_range = part.strip()
    if cell_range.startswith("'") and cell_range.endswith("'") and "!" in cell_range:
        cell_range = cell_range[1:-1]
    if "!" in cell_range:
        sheet, cell_range = cell_range.rsplit("!", 1)
        sheet_name = _normalize_sheet_name(sheet) or None
    return ExcelTargetRange(
        sheet_name=sheet_name,
        cell_range=_normalize_range(_strip_stray_quotes(cell_range.strip())),
    )


def _unquote_sheet_name(sheet: str) -> str:
    if len(sheet) >= 2 and sheet[0] == "'" and sheet[-1] == "'":
        return sheet[1:-1].replace("''", "'")
    return sheet


def _normalize_sheet_name(sheet: str) -> str:
    sheet = sheet.strip()
    unquoted = _unquote_sheet_name(sheet)
    if unquoted != sheet:
        return unquoted
    return _strip_stray_quotes(sheet)


def _strip_stray_quotes(value: str) -> str:
    return value.strip("'")


def _normalize_range(cell_range: str) -> str:
    match = re.fullmatch(r"([A-Z]+)(\d+):(\d+)", cell_range, flags=re.IGNORECASE)
    if match:
        col, start, end = match.groups()
        return f"{col.upper()}{start}:{col.upper()}{end}"
    return cell_range


def _resolve_sheet(
    sheetnames: list[str], requested: str | None, default_sheet: str | None
) -> str:
    return _resolve_sheets(sheetnames, requested, default_sheet)[0]


def _resolve_sheets(
    sheetnames: list[str], requested: str | None, default_sheet: str | None
) -> list[str]:
    if requested:
        if requested not in sheetnames:
            raise AssertionError(f"target sheet {requested!r} not found; available sheets: {sheetnames}")
        return [requested]
    if default_sheet:
        if default_sheet in sheetnames:
            return [default_sheet]
        candidates = [_normalize_sheet_name(part) for part in _split_range_list(default_sheet)]
        if candidates and all(candidate in sheetnames for candidate in candidates):
            return candidates
        raise AssertionError(
            f"default target sheet {default_sheet!r} not found; available sheets: {sheetnames}"
        )
    return [sheetnames[0]]


def _resolve_context_targets(
    sheetnames: list[str], targets: list[ExcelTargetRange], default_sheet: str | None
) -> tuple[list[tuple[str, ExcelTargetRange, bool]], list[str]]:
    resolved: list[tuple[str, ExcelTargetRange, bool]] = []
    missing: list[str] = []
    for target in targets:
        sheet_candidates = _context_sheet_candidates(sheetnames, target.sheet_name, default_sheet)
        for sheet_name in sheet_candidates:
            exists = sheet_name in sheetnames
            resolved.append((sheet_name, target, exists))
            if not exists and sheet_name not in missing:
                missing.append(sheet_name)
    return resolved, missing


def _context_sheet_candidates(
    sheetnames: list[str], requested: str | None, default_sheet: str | None
) -> list[str]:
    if requested:
        return [requested]
    if default_sheet:
        if default_sheet in sheetnames:
            return [default_sheet]
        candidates = [_normalize_sheet_name(part) for part in _split_range_list(default_sheet)]
        if candidates:
            return candidates
    return [sheetnames[0]]


def _coord_values(range_obj: Any) -> list[tuple[str, Any]]:
    if hasattr(range_obj, "value"):
        return [(range_obj.coordinate, range_obj.value)]
    out: list[tuple[str, Any]] = []
    for item in range_obj:
        if hasattr(item, "value"):
            out.append((item.coordinate, item.value))
        else:
            out.extend((cell.coordinate, cell.value) for cell in item)
    return out


def _formula_count(ws: Any) -> int:
    total = 0
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                total += 1
    return total


def _format_row(ws: Any, row: int, max_col: int, *, quote_strings: bool) -> str:
    openpyxl = _openpyxl()
    values = []
    for col in range(1, max_col + 1):
        letter = openpyxl.utils.get_column_letter(col)
        values.append(
            f"{letter}={_format_cell_value(ws.cell(row=row, column=col).value, quote_strings=quote_strings)}"
        )
    return " | ".join(values)


def _format_cell_value(value: Any, *, quote_strings: bool) -> str:
    if value is None:
        return "<blank>"
    if isinstance(value, str) and not quote_strings:
        return value
    return repr(value)


def _max_target_column(targets: list[ExcelTargetRange]) -> int:
    openpyxl = _openpyxl()
    max_col = 1
    for target in targets:
        try:
            _, _, end_col, _ = openpyxl.utils.cell.range_boundaries(target.cell_range)
        except ValueError:
            continue
        max_col = max(max_col, int(end_col))
    return max_col


def _openpyxl() -> Any:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - depends on optional environment
        raise ImportError(
            "Excel artifact helpers require openpyxl. Install it with `pip install openpyxl`."
        ) from exc
    return openpyxl
