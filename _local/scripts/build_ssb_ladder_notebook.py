"""Build a SpreadsheetBench Fabric notebook for the **Ladder** strategy.

Strategy L: cost-aware escalation ladder.
  Rung 0 (cheap):   gpt-4.1-mini + RLM(v6-custom) + excel_modify skill, max_turns=14
  Validator (free): escalate iff (not submitted) OR (no cells in answer_position
                    changed vs. the init workbook) OR (n_turns >= max).
  Rung 1 (strong):  gpt-5 parent + sub_lm=gpt-4.1-mini worker (S config), max_turns=14
                    only fires if rung 0 fails the validator.

This is the "smart bandit" for SSB: spend mini-only on questions mini handles
(detected for free via "did anything in the answer range actually change?") and
pay frontier price only on the residual.

Output records include rung_used, rung_costs, escalated boolean, and BOTH
attempts' token/turn metrics.

Usage:
  python scripts/build_ssb_ladder_notebook.py --run-id ssb-ladder-... [--smoke 5]
"""
import argparse, pathlib

ap = argparse.ArgumentParser()
ap.add_argument('--cheap-model', default='gpt-4.1-mini')
ap.add_argument('--strong-model', default='gpt-5')
ap.add_argument('--strong-sub-lm', default='gpt-4.1-mini')
ap.add_argument('--strong-effort', default='medium')
ap.add_argument('--run-id', required=True)
ap.add_argument('--smoke', type=int, default=0)
ap.add_argument('--out', default=None)
args = ap.parse_args()

WHEEL_PATH         = "/lakehouse/default/Files/fabric_rlm_longcot/wheels/fabric_rlm-0.2.1.dev2+excelskill-py3-none-any.whl"
DATASET_TAR        = "/lakehouse/default/Files/fabric_rlm_longcot/datasets/ssb_full_400.tar.gz"
DATASET_JSONL_NAME = "ssb_full_400.jsonl"
DATASET_HF_URL     = "https://huggingface.co/datasets/KAKA22/SpreadsheetBench/resolve/main/spreadsheetbench_verified_400.tar.gz"
RUN_ROOT           = f"/lakehouse/default/Files/fabric_rlm_adaptive_validation/spreadsheetbench/{args.run_id}"

cells = []
def cell(src):
    cells.append({'cell_type': 'code', 'metadata': {}, 'source': src.splitlines(keepends=True),
                  'outputs': [], 'execution_count': None})

cell(f"""%pip install -q dspy>=3.0.4 openpyxl
%pip install -q "{WHEEL_PATH}"
""")

cell(f"""import os, sys, json, time, tarfile, shutil, traceback, pathlib, urllib.request
import openpyxl
import fabric_rlm
from fabric_rlm import RLM, FabricLM

STRATEGY     = "L"
CHEAP_MODEL  = "{args.cheap_model}"
STRONG_MODEL = "{args.strong_model}"
STRONG_SUB   = "{args.strong_sub_lm}"
STRONG_EFF   = "{args.strong_effort}"
RUN_ID       = "{args.run_id}"
SMOKE_N      = {args.smoke}

DATASET_TAR        = "{DATASET_TAR}"
DATASET_JSONL_NAME = "{DATASET_JSONL_NAME}"
DATASET_HF_URL     = "{DATASET_HF_URL}"
RUN_ROOT           = pathlib.Path("{RUN_ROOT}")
RUN_ROOT.mkdir(parents=True, exist_ok=True)
RESULTS_PATH = RUN_ROOT / f"results_L.jsonl"
SUMMARY_PATH = RUN_ROOT / f"summary_L.json"
TRACES_DIR   = RUN_ROOT / f"traces_L"
TRACES_DIR.mkdir(exist_ok=True)

def stage(event, **kw):
    print(json.dumps({{"t": round(time.time(),2), "event": event, **kw}}, default=str))

stage("imports_done", fabric_rlm=getattr(fabric_rlm, '__version__', '?'))

WORK = pathlib.Path("/tmp/ssb_work"); WORK.mkdir(parents=True, exist_ok=True)
DS_DIR = WORK / "ds"

def _bootstrap_dataset():
    tar_p   = pathlib.Path(DATASET_TAR)
    jsonl_p = tar_p.parent / DATASET_JSONL_NAME
    if tar_p.exists() and jsonl_p.exists():
        stage("dataset_cache_hit", tar=str(tar_p)); return tar_p, jsonl_p
    stage("dataset_cache_miss_downloading_hf", url=DATASET_HF_URL)
    tar_p.parent.mkdir(parents=True, exist_ok=True)
    tmp_raw = WORK / "ssb_hf_raw.tar.gz"
    if not tmp_raw.exists():
        urllib.request.urlretrieve(DATASET_HF_URL, tmp_raw)
    raw_dir = WORK / "ssb_hf_raw"; raw_dir.mkdir(exist_ok=True)
    with tarfile.open(tmp_raw) as tf:
        tf.extractall(raw_dir)
    inner = next(raw_dir.glob("spreadsheetbench_verified_400*"), None) or raw_dir
    if not (inner / "dataset.json").exists():
        cands = list(raw_dir.rglob("dataset.json"))
        if cands: inner = cands[0].parent
    ds_json = json.load(open(inner / "dataset.json", encoding="utf-8"))
    spr_dir = inner / "spreadsheet"
    records = []
    for rec in ds_json:
        sid = str(rec["id"]); d = spr_dir / sid
        if not d.exists(): continue
        inits   = sorted(d.glob("*_init.xlsx"))
        goldens = sorted(d.glob("*_golden.xlsx"))
        if not (inits and goldens): continue
        prompt_p = d / "prompt.txt"
        instr = prompt_p.read_text(encoding="utf-8") if prompt_p.exists() else rec.get("instruction","")
        records.append({{"question_id": f"SSB_{{sid}}", "spreadsheet_id": sid,
                        "instruction": instr, "instruction_type": rec.get("instruction_type"),
                        "answer_position": rec["answer_position"], "answer_sheet": rec.get("answer_sheet"),
                        "init_file": inits[0].name, "golden_file": goldens[0].name}})
    with open(jsonl_p, "w", encoding="utf-8") as fh:
        for r in records: fh.write(json.dumps(r) + "\\n")
    stage_dir = WORK / "ssb_stage"
    if stage_dir.exists(): shutil.rmtree(stage_dir)
    stage_dir.mkdir()
    shutil.copytree(spr_dir, stage_dir / "spreadsheets")
    with tarfile.open(tar_p, "w:gz") as tf:
        tf.add(stage_dir / "spreadsheets", arcname="spreadsheets")
    stage("dataset_bootstrapped", n_records=len(records))
    return tar_p, jsonl_p

DATASET_TAR_P, DATASET_JSONL_P = _bootstrap_dataset()
if not DS_DIR.exists():
    DS_DIR.mkdir()
    with tarfile.open(DATASET_TAR_P) as tf:
        tf.extractall(DS_DIR)
n_xlsx = len(list(DS_DIR.rglob('*.xlsx')))
stage("dataset_extracted", path=str(DS_DIR), xlsx_count=n_xlsx)

records = [json.loads(l) for l in open(DATASET_JSONL_P, encoding='utf-8')]
if SMOKE_N > 0: records = records[:SMOKE_N]
stage("records_loaded", n=len(records))

def grade(out_xlsx, gold_xlsx, sheet, cell_range):
    try:
        wb_a = openpyxl.load_workbook(out_xlsx, data_only=True)
        wb_b = openpyxl.load_workbook(gold_xlsx, data_only=True)
    except Exception as e:
        return False, 0, 0, f"load_err: {{e}}"
    sname_a = sheet if (sheet and sheet in wb_a.sheetnames) else wb_a.sheetnames[0]
    sname_b = sheet if (sheet and sheet in wb_b.sheetnames) else wb_b.sheetnames[0]
    a = wb_a[sname_a]; b = wb_b[sname_b]
    try:
        rng_a = a[cell_range]; rng_b = b[cell_range]
    except Exception as e:
        return False, 0, 0, f"range_err: {{e}}"
    flat_a = [c.value for row in rng_a for c in row]
    flat_b = [c.value for row in rng_b for c in row]
    if len(flat_a) != len(flat_b):
        return False, 0, len(flat_b), f"len_mismatch {{len(flat_a)}} vs {{len(flat_b)}}"
    matches = 0
    for x, y in zip(flat_a, flat_b):
        if x == y: matches += 1
        else:
            try:
                if abs(float(x) - float(y)) < 1e-6: matches += 1; continue
            except Exception: pass
            sx = (str(x) if x is not None else "").strip()
            sy = (str(y) if y is not None else "").strip()
            if sx == sy: matches += 1
    return matches == len(flat_b), matches, len(flat_b), None

def cells_changed_in_range(init_path, work_path, sheet, cell_range):
    \"\"\"Free validator: did the model actually write into the target range?
    Returns (n_changed, n_total). n_changed==0 means the model didn't touch
    the answer area, which is a strong escalation signal.\"\"\"
    try:
        wb_i = openpyxl.load_workbook(init_path, data_only=True)
        wb_w = openpyxl.load_workbook(work_path, data_only=True)
    except Exception:
        return 0, 0
    si = sheet if (sheet and sheet in wb_i.sheetnames) else wb_i.sheetnames[0]
    sw = sheet if (sheet and sheet in wb_w.sheetnames) else wb_w.sheetnames[0]
    try:
        ri = wb_i[si][cell_range]; rw = wb_w[sw][cell_range]
    except Exception:
        return 0, 0
    fi = [c.value for row in ri for c in row]
    fw = [c.value for row in rw for c in row]
    n = len(fi); changed = sum(1 for x,y in zip(fi,fw) if x != y)
    return changed, n
""")

cell("""# === LADDER RUNNER ===
# Rung 0: gpt-4.1-mini + RLM (cheap)
# Validator (free + 1 cheap LM self-check call via FabricLM):
#   - escalate if rung 0 errored / didn't submit / hit max turns / didn't touch range
#   - escalate if cheap self-check (gpt-4.1-mini) judges output 'NO'
#   - if self-check itself errors (auth, etc.), DO NOT force-escalate — trust rung 0
# Rung 1: gpt-5 parent + sub_lm=gpt-4.1-mini  (only if validator escalates)

import dspy
_check_lm = FabricLM(model=CHEAP_MODEL, temperature=0.0)
dspy.configure(lm=_check_lm)

class SsbSelfCheck(dspy.Signature):
    \"\"\"Given an Excel manipulation instruction and the actual cell values
    produced by an automated solver in the target answer range, judge whether
    those values look correct for the instruction. Be strict but fair: say NO
    only if values look clearly wrong (all zeros, all duplicates, obvious type
    mismatch, missing data). If values look plausible for the instruction, say YES.\"\"\"
    instruction: str = dspy.InputField()
    answer_sheet: str = dspy.InputField()
    answer_range: str = dspy.InputField()
    cell_values: str = dspy.InputField(desc="Flat list of values produced in the answer range")
    looks_correct: str = dspy.OutputField(desc="exactly 'YES' or 'NO'")
    reason: str = dspy.OutputField(desc="one short sentence")

_self_check = dspy.Predict(SsbSelfCheck)

def get_cell_values_in_range(xlsx_path, sheet, cell_range, max_items=100):
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception:
        return [], None
    sn = sheet if (sheet and sheet in wb.sheetnames) else wb.sheetnames[0]
    try:
        rng = wb[sn][cell_range]
    except Exception:
        return [], sn
    flat = [c.value for row in rng for c in row]
    return flat[:max_items], sn

def build_rlm(model_name, *, sub_lm=None, effort=None):
    lm_kwargs = dict(temperature=1.0)
    if effort and model_name.startswith(('gpt-5','o3','o1')):
        lm_kwargs['reasoning_effort'] = effort
    base_lm = FabricLM(model=model_name, **lm_kwargs)
    rlm_kwargs = dict(engine='v6-custom', max_turns=14,
                      skills=['excel_modify'], timeout=300.0)
    if sub_lm:
        rlm_kwargs['sub_lm'] = sub_lm
    return RLM(signature='question -> answer', lm=base_lm, **rlm_kwargs)

t_start = time.time(); n_pass = 0; n_escalated = 0
with RESULTS_PATH.open('w', encoding='utf-8') as out_fh:
    for idx, rec in enumerate(records):
        qid = rec['question_id']; sid = str(rec['spreadsheet_id'])
        init_src = DS_DIR / 'spreadsheets' / sid / rec['init_file']
        gold_src = DS_DIR / 'spreadsheets' / sid / rec['golden_file']
        sheet = rec.get('answer_sheet') or ''
        ans_pos = rec['answer_position']
        instr = rec['instruction']
        work_dir = WORK / qid; work_dir.mkdir(exist_ok=True)
        rec_out = {'strategy': 'L', 'question_id': qid, 'spreadsheet_id': sid,
                   'instruction_type': rec.get('instruction_type'),
                   'answer_sheet': sheet, 'answer_position': ans_pos}
        attempts = []
        try:
            # ---- RUNG 0 (cheap) ----
            work_xlsx_0 = work_dir / 'work_rung0.xlsx'
            shutil.copyfile(init_src, work_xlsx_0)
            prompt = (f"You are given an Excel workbook at {work_xlsx_0}.\\n"
                      f"Instruction: {instr}\\n"
                      f"The grader reads cells with openpyxl(data_only=True), so write "
                      f"COMPUTED VALUES (not Excel formulas) into the workbook and SAVE "
                      f"it back to the same path. Target answer range: sheet={sheet!r}, "
                      f"cells={ans_pos!r}.")
            t0 = time.perf_counter()
            rlm0 = build_rlm(CHEAP_MODEL)
            res0 = rlm0.run({'question': prompt})
            elapsed0 = time.perf_counter() - t0
            traj0 = res0.trajectory
            turns0 = list(getattr(traj0, 'turns', []) or []) if traj0 is not None else []
            p0 = sum((getattr(t,'prompt_tokens',None) or 0) for t in turns0)
            c0 = sum((getattr(t,'completion_tokens',None) or 0) for t in turns0)
            changed0, totalc = cells_changed_in_range(init_src, work_xlsx_0, sheet, ans_pos)
            submitted0 = bool(getattr(res0, 'submitted', False))
            # Self-check call (cheap, ~$0.001/Q)
            sc_says = ''
            sc_reason = ''
            sc_p = sc_c = 0
            try:
                vals, snused = get_cell_values_in_range(work_xlsx_0, sheet, ans_pos)
                vals_str = json.dumps(vals, default=str)[:1500]
                sc = _self_check(instruction=instr, answer_sheet=(snused or sheet or ''),
                                 answer_range=ans_pos, cell_values=vals_str)
                sc_says = (sc.looks_correct or '').strip().upper()
                sc_reason = (sc.reason or '')[:200]
                hist = getattr(_check_lm, 'history', []) or []
                if hist:
                    u = (hist[-1].get('usage') or {}) if isinstance(hist[-1], dict) else {}
                    sc_p = u.get('prompt_tokens') or u.get('input_tokens') or 0
                    sc_c = u.get('completion_tokens') or u.get('output_tokens') or 0
            except Exception as e:
                sc_says = 'ERR'; sc_reason = repr(e)[:200]
            attempts.append({'rung': 0, 'model': CHEAP_MODEL, 'sub_lm': None,
                             'submitted': submitted0, 'n_turns': len(turns0),
                             'cells_changed_in_range': changed0,
                             'cells_in_range_total': totalc,
                             'prompt_tokens': p0 + sc_p,
                             'completion_tokens': c0 + sc_c,
                             'elapsed_seconds': round(elapsed0, 2),
                             'self_check': sc_says,
                             'self_check_reason': sc_reason})
            # Validator: escalate on hard failure signals OR explicit NO from self-check.
            # If self-check ERR'd, we trust rung 0 (don't force-escalate).
            escalate = (
                (not submitted0) or
                (changed0 == 0) or
                (len(turns0) >= 14) or
                (sc_says == 'NO')
            )
            stage('rung0_done', qid=qid, submitted=submitted0,
                  changed=f'{changed0}/{totalc}', turns=len(turns0),
                  self_check=sc_says, escalate=escalate)
            chosen_xlsx = work_xlsx_0
            # ---- RUNG 1 (strong) — only if validator escalates ----
            if escalate:
                n_escalated += 1
                work_xlsx_1 = work_dir / 'work_rung1.xlsx'
                shutil.copyfile(init_src, work_xlsx_1)
                prompt1 = (f"You are given an Excel workbook at {work_xlsx_1}.\\n"
                           f"Instruction: {instr}\\n"
                           f"The grader reads cells with openpyxl(data_only=True), so write "
                           f"COMPUTED VALUES (not Excel formulas). Target range: "
                           f"sheet={sheet!r}, cells={ans_pos!r}.")
                t1 = time.perf_counter()
                rlm1 = build_rlm(STRONG_MODEL, sub_lm=STRONG_SUB, effort=STRONG_EFF)
                res1 = rlm1.run({'question': prompt1})
                elapsed1 = time.perf_counter() - t1
                traj1 = res1.trajectory
                turns1 = list(getattr(traj1, 'turns', []) or []) if traj1 is not None else []
                p1 = sum((getattr(t,'prompt_tokens',None) or 0) for t in turns1)
                c1 = sum((getattr(t,'completion_tokens',None) or 0) for t in turns1)
                submitted1 = bool(getattr(res1, 'submitted', False))
                attempts.append({'rung': 1, 'model': STRONG_MODEL, 'sub_lm': STRONG_SUB,
                                 'submitted': submitted1, 'n_turns': len(turns1),
                                 'prompt_tokens': p1, 'completion_tokens': c1,
                                 'elapsed_seconds': round(elapsed1, 2)})
                chosen_xlsx = work_xlsx_1
                stage('rung1_done', qid=qid, submitted=submitted1, turns=len(turns1))
            # Grade the chosen attempt
            passed, m, n, gerr = grade(str(chosen_xlsx), str(gold_src), sheet, ans_pos)
            (TRACES_DIR / f'trace_{qid}.json').write_text(json.dumps({
                'qid': qid, 'instruction': instr, 'attempts': attempts,
                'passed': passed, 'rung_used': attempts[-1]['rung'],
                'escalated': len(attempts) > 1,
            }, default=str, indent=2), encoding='utf-8')
            rec_out.update({
                'passed': passed, 'cells_matched': m, 'cells_total': n, 'grade_err': gerr,
                'rung_used': attempts[-1]['rung'], 'escalated': len(attempts) > 1,
                'attempts': attempts,
                'total_prompt_tokens':     sum(a['prompt_tokens'] for a in attempts),
                'total_completion_tokens': sum(a['completion_tokens'] for a in attempts),
                'total_elapsed_seconds':   sum(a['elapsed_seconds'] for a in attempts),
            })
        except Exception as e:
            rec_out.update({'passed': False, 'error': repr(e),
                            'traceback': traceback.format_exc()[:500],
                            'attempts': attempts})
        if rec_out.get('passed'): n_pass += 1
        out_fh.write(json.dumps(rec_out, default=str) + '\\n'); out_fh.flush()
        stage('q_done', idx=idx+1, qid=qid, passed=rec_out.get('passed'),
              cells=f"{rec_out.get('cells_matched',0)}/{rec_out.get('cells_total',0)}",
              rung=rec_out.get('rung_used'), escalated=rec_out.get('escalated'))

summary = {'strategy': 'L', 'cheap_model': CHEAP_MODEL,
           'strong_model': STRONG_MODEL, 'strong_sub_lm': STRONG_SUB,
           'strong_effort': STRONG_EFF, 'run_id': RUN_ID,
           'n': len(records), 'n_passed': n_pass,
           'n_escalated': n_escalated,
           'escalation_rate': round(n_escalated / max(1, len(records)), 4),
           'pass_rate': round(n_pass / max(1, len(records)), 4),
           'total_seconds': round(time.time() - t_start, 1)}
SUMMARY_PATH.write_text(json.dumps(summary, indent=2), encoding='utf-8')
print('\\n=== SUMMARY ===')
print(json.dumps(summary, indent=2))
""")

nb = {
    'nbformat': 4, 'nbformat_minor': 5,
    'metadata': {
        'kernelspec': {'display_name': 'Python 3.12', 'language': 'python', 'name': 'python3.12'},
        'language_info': {'name': 'python', 'version': '3.12'},
        'kernel_info': {'name': 'jupyter', 'jupyter_kernel_name': 'python3.12'},
        'microsoft': {'language': 'python', 'language_group': 'jupyter_python'},
        'dependencies': {'lakehouse': {
            'default_lakehouse': '9d10bce5-1edc-4875-83c4-ac0a98a02775',
            'default_lakehouse_name': 'diagnostic',
            'default_lakehouse_workspace_id': '82ad2591-974a-4ad4-ace6-e24879274a4b'}},
    },
    'cells': cells,
}

out = pathlib.Path(args.out) if args.out else pathlib.Path("notebooks") / f"ssb_L_ladder.ipynb"
out.parent.mkdir(parents=True, exist_ok=True)
import json as _json
out.write_text(_json.dumps(nb, indent=1), encoding='utf-8')
print(f"wrote {out}  cells={len(cells)}  strategy=L  cheap={args.cheap_model}  strong={args.strong_model} sub_lm={args.strong_sub_lm}")
